from openpyxl import load_workbook

filename="HelloWorld.xlsx"
workbook = load_workbook(filename=filename)
print(workbook.sheetnames)
sheet = workbook.active
print("active sheet title ",sheet.title)

print("max row which have data - ", sheet.max_row )
print("max column which have data - ", sheet.max_column )

cell11 = sheet.cell(row=1, column=1)
print("data in cell 1,1 - ",cell11.value)

cell_a2 = sheet["A2"]
print("data in cell A2 -",cell_a2.value)

all_cell_a_column = sheet["A"]
print("printing all A column values")
for cell in all_cell_a_column:
    print(cell.value)


all_cell_b_column = sheet["B"]
print("printing all B column values")
for cell in all_cell_b_column:
    print(cell.value)

print("printing all row and column values")
for row_value in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1,max_col=sheet.max_column, values_only=True):
    print(row_value)

print("printing all column and column values")
for column_value in sheet.iter_cols(min_row=1, max_row=sheet.max_row, min_col=1,max_col=sheet.max_column, values_only=True):
    print(column_value)

print(f"{filename} reading done !")
