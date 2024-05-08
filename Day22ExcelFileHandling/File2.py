from openpyxl import Workbook
filename = "HelloWorld.xlsx"
workbook = Workbook()
spreadsheet = workbook.active
print("spreadsheet title ", spreadsheet.title)

spreadsheet["A1"] = "Hello"
spreadsheet["B1"] = "World!"

c1 = spreadsheet.cell(row=2, column=1)
c1.value = "Welcome"
c2 = spreadsheet.cell(row=2, column=2)
c2.value = "Everyone"

workbook.save(filename=filename)
print(f"{filename} created successfully")
